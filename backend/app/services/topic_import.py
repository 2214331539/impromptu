import csv
import json
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import httpx
from pydantic import BaseModel, Field, ValidationError

from app.core.config import settings
from app.core.exceptions import AppError
from app.models.entities import Difficulty
from app.schemas.models import TopicImportItem, TopicImportPreviewOut


class _AIImportResult(BaseModel):
    topics: list[str] = Field(min_length=1)
    warnings: list[str] = []


class TopicImportService:
    allowed_extensions = {".txt", ".csv", ".tsv", ".xlsx"}

    def preview(
        self,
        *,
        name: str | None,
        description: str | None,
        raw_text: str | None,
        filename: str | None,
        file_content: bytes | None,
    ) -> TopicImportPreviewOut:
        source = self._source_text(raw_text, filename, file_content)
        if len(source) < 2:
            raise AppError("EMPTY_IMPORT_SOURCE", "Import content is too short.", 422)
        if not settings.ai_import_configured:
            raise AppError("AI_IMPORT_NOT_CONFIGURED", "AI import is not configured.", 503)

        result = self._request_ai(
            source=source,
            requested_name=(name or "").strip(),
            requested_description=(description or "").strip(),
        )
        max_topics = max(1, settings.ai_import_max_topics)
        topics = self._dedupe_topics(result.topics)[:max_topics]
        if not topics:
            raise AppError("AI_IMPORT_EMPTY", "AI did not return usable topics.", 422)
        warnings = list(result.warnings)
        if len(result.topics) > max_topics:
            warnings.append(f"Only the first {max_topics} topics were kept.")
        return TopicImportPreviewOut(
            name=(name or "AI Imported Topics").strip(),
            description=(description or "").strip(),
            topics=topics,
            warnings=warnings,
        )

    def _source_text(
        self,
        raw_text: str | None,
        filename: str | None,
        file_content: bytes | None,
    ) -> str:
        chunks = []
        if raw_text and raw_text.strip():
            chunks.append(raw_text.strip())
        if file_content:
            chunks.append(self._read_file(filename or "", file_content))
        return "\n\n".join(item for item in chunks if item.strip()).strip()

    def _read_file(self, filename: str, content: bytes) -> str:
        if len(content) > 2 * 1024 * 1024:
            raise AppError("IMPORT_FILE_TOO_LARGE", "Import file must be 2MB or smaller.", 413)
        suffix = Path(filename).suffix.lower()
        if suffix not in self.allowed_extensions:
            raise AppError("IMPORT_FILE_TYPE", "Only txt, csv, tsv, and xlsx files are supported.", 415)
        if suffix == ".xlsx":
            return self._read_xlsx(content)
        text = self._decode_text(content)
        if suffix in {".csv", ".tsv"}:
            delimiter = "\t" if suffix == ".tsv" else ","
            rows = csv.reader(StringIO(text), delimiter=delimiter)
            return "\n".join(" | ".join(cell.strip() for cell in row if cell.strip()) for row in rows)
        return text

    @staticmethod
    def _decode_text(content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise AppError("IMPORT_FILE_ENCODING", "Could not decode the import file.", 422)

    @staticmethod
    def _read_xlsx(content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise AppError("XLSX_READER_MISSING", "Excel import dependency is not installed.", 500) from exc
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if cells:
                    lines.append(" | ".join(cells))
        workbook.close()
        return "\n".join(lines)

    def _request_ai(self, *, source: str, requested_name: str, requested_description: str) -> _AIImportResult:
        url = settings.openai_base_url.rstrip("/") + "/chat/completions"
        payload = {
            "model": settings.openai_model,
            "temperature": 0.2,
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "You clean teacher-provided topic terms for an impromptu speaking topic bank. "
                        "The input is usually many nouns or short academic phrases copied from Excel. "
                        "Return strict JSON only with this schema: "
                        '{"topics": ["Growth Mindset", "Cognitive Dissonance"], "warnings": string[]}. '
                        "Do not create a topic bank name, description, categories, tags, questions, prompts, "
                        "or explanations. Do not rewrite, expand, translate, or polish the topic terms. "
                        "Only remove numbering, empty rows, table headers, and duplicates while preserving "
                        "the original order and wording as much as possible."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(
                        {
                            "source": source[:20000],
                            "max_topics": settings.ai_import_max_topics,
                        },
                        ensure_ascii=False,
                    ),
                },
            ],
        }
        try:
            with httpx.Client(timeout=settings.ai_import_timeout_seconds) as client:
                response = client.post(
                    url,
                    headers={"Authorization": f"Bearer {settings.openai_api_key}"},
                    json=payload,
                )
                response.raise_for_status()
        except httpx.HTTPStatusError as exc:
            raise AppError("AI_IMPORT_FAILED", f"AI provider returned {exc.response.status_code}.", 502) from exc
        except httpx.HTTPError as exc:
            raise AppError("AI_IMPORT_FAILED", "AI provider request failed.", 502) from exc

        try:
            data = response.json()
            content = data["choices"][0]["message"]["content"]
            parsed = self._parse_json_content(content)
            return _AIImportResult.model_validate(parsed)
        except (KeyError, TypeError, ValueError, ValidationError) as exc:
            raise AppError("AI_IMPORT_INVALID_RESPONSE", "AI response could not be parsed.", 502) from exc

    @staticmethod
    def _parse_json_content(content: str | dict[str, Any]) -> dict[str, Any]:
        if isinstance(content, dict):
            return content
        text = content.strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        if not text.startswith("{"):
            start = text.find("{")
            end = text.rfind("}")
            if start >= 0 and end > start:
                text = text[start : end + 1]
        return json.loads(text)

    @staticmethod
    def _dedupe_topics(topics: list[str]) -> list[TopicImportItem]:
        seen: set[str] = set()
        result: list[TopicImportItem] = []
        for topic in topics:
            prompt = str(topic).strip().strip("-•*0123456789.、) ")
            if len(prompt) < 2:
                continue
            normalized_prompt = " ".join(prompt.lower().split())
            if normalized_prompt in seen:
                continue
            seen.add(normalized_prompt)
            result.append(
                TopicImportItem(
                    prompt=prompt[:2000],
                    category="Topic",
                    difficulty=Difficulty.MEDIUM,
                    tags="",
                )
            )
        return result
